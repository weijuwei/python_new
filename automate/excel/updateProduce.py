# <<Python编程快速上手-让繁琐工作自动化>>第十二章
# 更新商品单价项目

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

ws = wb.active

# 将字典中包含商品的单价进行修改
PRICE_UPDATES = {'Garlic': 3.07,
				 'Celery': 1.19,
				 'Lemon': 1.27}

for row in range(2,ws.max_row):  # 对商品名字列进行遍历
	productName = ws['A' + str(row)].value

	if productName in PRICE_UPDATES:
		ws['B'+str(row)].value = PRICE_UPDATES[productName]
		# print(productName + '---->' + str(ws['B'+str(row)].value))

wb.save('produceSales_v2.xlsx')  # 写到文件中