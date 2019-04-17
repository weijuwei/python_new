# <<Python编程快速上手-让繁琐工作自动化>>第十二章
# 从excel中读取数据

import openpyxl

print("Opening workbook.....")

wb = openpyxl.load_workbook('censuspopdata.xlsx')
ws = wb.active

countyData = {}

for row in range(2,ws.max_row+1):
	state = ws['B' + str(row)].value
	county = ws['C' + str(row)].value
	pop = ws['D' + str(row)].value

	countyData.setdefault(state,{})
	countyData[state].setdefault(county,{'tracts':0,'pop':0})

	countyData[state][county]['tracts'] += 1
	countyData[state][county]['pop'] += int(pop)

print(countyData['WY']['Weston'])