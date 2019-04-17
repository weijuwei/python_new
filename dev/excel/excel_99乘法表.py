import openpyxl,os

wb = openpyxl.Workbook()

ws = wb.active
ws.title = '99乘法表'

num = 9

for i in range(2,11):
	ws.cell(row=1,column=i).value = i - 1
	ws.cell(row=i,column=1).value = i - 1

for i in range(2,11):
	# r = ws.cell(row=1,column=i).value
	# c = ws.cell(row=1,column=i).value
	# ws.cell(row=i,column=i).value = r * c
	# print(r,c)
	for j in range(2,11):
		row = ws.cell(row=i,column=1).value
		column = ws.cell(row=1,column=j).value
		ws.cell(row=i,column=j).value = row * column

if not os.path.isfile('99乘法表.xlsx')
	wb.save('99乘法表.xlsx')
